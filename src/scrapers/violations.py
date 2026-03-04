"""CPSC Import Violations (Notice of Violation) data scraper."""

import logging
import re
from pathlib import Path

import httpx

from ..normalization.manufacturers import normalize_manufacturer, extract_fiscal_year
from ..validation.quality import score_import_violation

logger = logging.getLogger(__name__)

# The URL may change with date — try multiple patterns
NOV_URLS = [
    "https://www.cpsc.gov/s3fs-public/CPSC-NOV-DATA-2026-02-19.xlsx",
    "https://www.cpsc.gov/s3fs-public/CPSC-NOV-DATA-2025-12-31.xlsx",
    "https://www.cpsc.gov/s3fs-public/CPSC-NOV-DATA-2025-06-30.xlsx",
]

RAW_DIR = Path(__file__).parent.parent.parent / "data" / "raw"

# Column name mapping
COLUMN_MAP = {
    "nov sent": "nov_date",
    "product name": "product_name",
    "model number": "model_number",
    "model no": "model_number",
    "model no.": "model_number",
    "sample#": "sample_number",
    "sample #": "sample_number",
    "sample number": "sample_number",
    "requested domestic action": "domestic_action",
    "domestic action": "domestic_action",
    "requested cbp action": "cbp_action",
    "cbp action": "cbp_action",
    "viol": "violation_type",
    "violation": "violation_type",
    "violation type": "violation_type",
    "cit": "citation",
    "citation": "citation",
    "firm name": "firm_name",
    "firm address": "firm_address",
    "firm city": "firm_city",
    "country": "country",
}


def download_violations() -> str | None:
    """Download the import violations Excel file. Returns path or None."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    cached = RAW_DIR / "cpsc_nov_data.xlsx"
    if cached.exists():
        logger.info(f"Using cached violations file: {cached}")
        return str(cached)

    for url in NOV_URLS:
        logger.info(f"Trying violations URL: {url}")
        try:
            resp = httpx.get(url, timeout=120, follow_redirects=True,
                           headers={"User-Agent": "CPSC-Product-Safety-Tracker/1.0 (nathanmauricegoldberg@gmail.com)"})
            if resp.status_code == 200:
                cached.write_bytes(resp.content)
                logger.info(f"Downloaded {len(resp.content):,} bytes")
                return str(cached)
            else:
                logger.warning(f"  HTTP {resp.status_code}")
        except Exception as e:
            logger.warning(f"  Failed: {e}")

    logger.warning("Could not download import violations file")
    return None


def parse_violations(xlsx_path: str) -> list[dict]:
    """Parse import violations from the Excel file."""
    import openpyxl

    records = []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row
        header_row = None
        header_idx = 0
        for i, row in enumerate(rows):
            if row and any(str(c).strip().lower() in COLUMN_MAP for c in row if c):
                header_row = row
                header_idx = i
                break

        if not header_row:
            continue

        # Map columns
        col_map = {}
        for j, cell in enumerate(header_row):
            if cell:
                key = str(cell).strip().lower()
                if key in COLUMN_MAP:
                    col_map[j] = COLUMN_MAP[key]

        if not col_map:
            continue

        # Parse data rows
        for row in rows[header_idx + 1:]:
            if not row or all(c is None for c in row):
                continue

            mapped = {}
            for j, field_name in col_map.items():
                if j < len(row):
                    val = row[j]
                    if val is not None:
                        mapped[field_name] = str(val).strip()
                    else:
                        mapped[field_name] = ""
                else:
                    mapped[field_name] = ""

            firm_name = mapped.get("firm_name", "")
            if not firm_name:
                continue

            # Normalize date
            nov_date = mapped.get("nov_date", "")
            if nov_date:
                # Handle datetime objects from openpyxl
                if hasattr(nov_date, "strftime"):
                    nov_date = nov_date.strftime("%Y-%m-%d")
                elif re.match(r"\d{4}-\d{2}-\d{2}", nov_date):
                    pass  # already formatted
                else:
                    # Try to clean up
                    nov_date = str(nov_date).split(" ")[0] if " " in str(nov_date) else str(nov_date)

            normalized = normalize_manufacturer(firm_name)
            fiscal_year = extract_fiscal_year(nov_date)

            record = {
                "nov_date": nov_date,
                "product_name": mapped.get("product_name", ""),
                "model_number": mapped.get("model_number", ""),
                "sample_number": mapped.get("sample_number", ""),
                "domestic_action": mapped.get("domestic_action", ""),
                "cbp_action": mapped.get("cbp_action", ""),
                "violation_type": mapped.get("violation_type", ""),
                "citation": mapped.get("citation", ""),
                "firm_name": firm_name,
                "firm_address": mapped.get("firm_address", ""),
                "firm_city": mapped.get("firm_city", ""),
                "country": mapped.get("country", ""),
                "normalized_firm": normalized,
                "fiscal_year": fiscal_year,
            }
            record["quality_score"] = score_import_violation(record)
            records.append(record)

    wb.close()
    logger.info(f"Parsed {len(records)} import violations from {len(wb.sheetnames)} sheets")
    return records
